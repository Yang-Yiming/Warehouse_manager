import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import datetime
import openpyxl
import os
import json

class WarehouseManager:
    def __init__(self, root):
        self.root = root
        self.root.title('仓库物资管理系统')
        self.data = []  # 存储物资操作信息的列表
        self.inventory = {}  # 存储当前库存信息，格式: {物资编号: {物品信息}}
        self.current_view = 'operations'  # 当前视图模式：'operations'或'inventory'
        
        # 初始化路径
        self.init_paths()
        
        # 加载配置
        self.load_config()
        
        # 加载数据
        self.load_data()
        self.load_inventory()
        
        # 创建界面
        self.create_widgets()
        
    def init_paths(self):
        """初始化路径设置"""
        self.data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
        self.output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
        self.data_file = os.path.join(self.data_dir, 'warehouse_data.json')
        self.inventory_file = os.path.join(self.data_dir, 'inventory_data.json')
        self.config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
        
        # 确保目录存在
        for directory in [self.data_dir, self.output_dir]:
            if not os.path.exists(directory):
                os.makedirs(directory)
                
    def load_config(self):
        """加载配置文件"""
        self.organizations = []
        self.operators = []
        
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.organizations = config.get('organization', {}).get('val', [])
                    self.operators = config.get('operators', {}).get('val', [])
            except Exception as e:
                messagebox.showerror('配置加载错误', f'无法加载配置: {str(e)}')
                
    def load_data(self):
        """从文件加载操作数据"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    old_data = json.load(f)
                
                # 转换旧数据到新格式
                self.data = []
                for item in old_data:
                    new_item = {
                        "提交时间": item.get("提交时间", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                        "物资编号": item.get('物资编号', ''),
                        "物品名称": item.get('物品名称', ''),
                        "物资操作": item.get('物资操作', '入库'),  # 默认为入库
                        "所属组织": item.get('所属组织', ''),
                        "物品数量": item.get('物品数量', 0),
                        "时间": item.get('时间', ''),
                        "操作人": item.get('操作人', ''),
                        "提交者": item.get('提交者', '')
                    }
                    self.data.append(new_item)
                
                self.save_data()  # 保存转换后的数据
            except Exception as e:
                messagebox.showerror('数据加载错误', f'无法加载数据: {str(e)}')
    
    def load_inventory(self):
        """从文件加载库存数据"""
        if os.path.exists(self.inventory_file):
            try:
                with open(self.inventory_file, 'r', encoding='utf-8') as f:
                    self.inventory = json.load(f)
            except Exception as e:
                messagebox.showerror('库存数据加载错误', f'无法加载库存数据: {str(e)}')
                self.inventory = {}
        else:
            # 如果库存文件不存在，根据操作记录重新生成库存
            self.rebuild_inventory_from_operations()
    
    def rebuild_inventory_from_operations(self):
        """根据操作记录重建库存数据"""
        self.inventory = {}
        
        for item in self.data:
            item_id = item.get('物资编号', '')
            operation = item.get('物资操作', '')
            qty = item.get('物品数量', 0)
            
            if not item_id:
                continue
                
            if operation == '入库' or operation == '物资增添':
                if item_id not in self.inventory:
                    # 新物品，添加到库存
                    self.inventory[item_id] = {
                        "物资编号": item_id,
                        "物品名称": item.get('物品名称', ''),
                        "所属组织": item.get('所属组织', ''),
                        "物品数量": qty,
                        "最后操作": item.get('物资操作', ''),
                        "最后操作人": item.get('操作人', ''),
                        "最后操作时间": item.get('时间', ''),
                        "备注": ""
                    }
                else:
                    # 现有物品，增加数量
                    self.inventory[item_id]['物品数量'] += qty
                    self.inventory[item_id]['最后操作'] = operation
                    self.inventory[item_id]['最后操作人'] = item.get('操作人', '')
                    self.inventory[item_id]['最后操作时间'] = item.get('时间', '')
            
            elif operation == '出库':
                # 完全出库，从库存中移除
                if item_id in self.inventory:
                    del self.inventory[item_id]
            
            elif operation == '部分出库':
                # 部分出库，减少数量
                if item_id in self.inventory:
                    self.inventory[item_id]['物品数量'] -= qty
                    if self.inventory[item_id]['物品数量'] <= 0:
                        # 如果数量减至0或以下，移除物品
                        del self.inventory[item_id]
                    else:
                        # 更新最后操作信息
                        self.inventory[item_id]['最后操作'] = operation
                        self.inventory[item_id]['最后操作人'] = item.get('操作人', '')
                        self.inventory[item_id]['最后操作时间'] = item.get('时间', '')
        
        # 保存重建后的库存
        self.save_inventory()
    
    def save_data(self):
        """保存数据到文件"""
        try:
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror('数据保存错误', f'无法保存数据: {str(e)}')

    def save_inventory(self):
        """保存库存数据到文件"""
        try:
            with open(self.inventory_file, 'w', encoding='utf-8') as f:
                json.dump(self.inventory, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror('库存数据保存错误', f'无法保存库存数据: {str(e)}')

    def create_widgets(self):
        """创建界面组件"""
        self.create_view_selector()
        self.create_search_panel()
        self.create_button_panel()
        self.create_table()
        self.update_table()
    
    def create_view_selector(self):
        """创建视图选择器"""
        selector_frame = tk.Frame(self.root)
        selector_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.view_var = tk.StringVar(value='operations')
        
        tk.Radiobutton(selector_frame, text='操作记录', variable=self.view_var, 
                      value='operations', command=self.switch_view).pack(side=tk.LEFT)
        tk.Radiobutton(selector_frame, text='当前库存', variable=self.view_var, 
                      value='inventory', command=self.switch_view).pack(side=tk.LEFT, padx=10)
    
    def switch_view(self):
        """切换视图模式"""
        self.current_view = self.view_var.get()
        # 切换视图时重新创建表格
        self.create_table()
        self.update_table()
    
    def create_search_panel(self):
        """创建搜索面板"""
        search_frame = tk.Frame(self.root)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Label(search_frame, text='搜索:').pack(side=tk.LEFT)
        
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=tk.LEFT, padx=5)
        search_entry.bind('<KeyRelease>', lambda e: self.update_table())

    def create_button_panel(self):
        """创建按钮面板"""
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Button(btn_frame, text='入库', command=self.add_item).pack(side=tk.LEFT)
        tk.Button(btn_frame, text='出库', command=lambda: self.remove_item('出库')).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text='物资增添', command=lambda: self.add_quantity('物资增添')).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text='部分出库', command=lambda: self.remove_item('部分出库')).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text='导入Excel', command=self.import_excel).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text='导出Excel', command=self.export_excel).pack(side=tk.LEFT, padx=5)
        # 添加重建库存按钮
        tk.Button(btn_frame, text='重建库存', command=self.rebuild_inventory).pack(side=tk.LEFT, padx=5)

    def rebuild_inventory(self):
        """手动重建库存数据"""
        if messagebox.askyesno('确认', '确定要根据所有操作记录重新构建库存数据吗？'):
            self.rebuild_inventory_from_operations()
            self.update_table()
            messagebox.showinfo('完成', '库存数据已重建')

    def create_table(self):
        """创建数据表格"""
        # 移除现有表格（如果存在）
        if hasattr(self, 'tree'):
            self.tree.destroy()
        
        if self.current_view == 'operations':
            # 操作记录视图
            columns = ('物资编号', '物品名称', '物资操作', '所属组织', '物品数量', '时间', '操作人', '提交者', '提交时间')
        else:
            # 库存视图
            columns = ('物资编号', '物品名称', '所属组织', '物品数量', '最后操作', '最后操作人', '最后操作时间', '备注')
        
        self.tree = ttk.Treeview(self.root, columns=columns, show='headings')
        
        # 配置列
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by(c, False))
            if col in ['提交时间', '操作人', '提交者', '最后操作时间', '最后操作人', '备注']:
                self.tree.column(col, width=150)
            else:
                self.tree.column(col, width=100)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
    def update_table(self):
        """更新表格数据显示"""
        search = self.search_var.get().lower()
        
        # 清空表格
        for row in self.tree.get_children():
            self.tree.delete(row)
            
        if self.current_view == 'operations':
            # 更新操作记录视图
            for item in self.data:
                # 检查是否符合搜索条件
                searchable_fields = [
                    str(item.get('物资编号', '')), 
                    item.get('物品名称', ''), 
                    item.get('物资操作', ''),
                    item.get('所属组织', ''), 
                    str(item.get('物品数量', '')),
                    item.get('时间', ''),
                    item.get('操作人', ''),
                    item.get('提交者', ''),
                    item.get('提交时间', '')
                ]
                
                if search and not any(search in field.lower() for field in searchable_fields):
                    continue
                    
                self.tree.insert('', tk.END, values=(
                    item.get('物资编号', ''),
                    item.get('物品名称', ''),
                    item.get('物资操作', ''),
                    item.get('所属组织', ''),
                    item.get('物品数量', 0),
                    item.get('时间', ''),
                    item.get('操作人', ''),
                    item.get('提交者', ''),
                    item.get('提交时间', '')
                ))
        else:
            # 更新库存视图
            for item_id, item in self.inventory.items():
                # 检查是否符合搜索条件
                searchable_fields = [
                    item_id,
                    item.get('物品名称', ''),
                    item.get('所属组织', ''),
                    str(item.get('物品数量', '')),
                    item.get('最后操作', ''),
                    item.get('最后操作人', ''),
                    item.get('最后操作时间', ''),
                    item.get('备注', '')
                ]
                
                if search and not any(search in field.lower() for field in searchable_fields):
                    continue
                    
                self.tree.insert('', tk.END, values=(
                    item_id,
                    item.get('物品名称', ''),
                    item.get('所属组织', ''),
                    item.get('物品数量', 0),
                    item.get('最后操作', ''),
                    item.get('最后操作人', ''),
                    item.get('最后操作时间', ''),
                    item.get('备注', '')
                ))
    
    def generate_new_id(self):
        """检查物资编号是否重复，不再自动生成"""
        return None  # 返回None表示不自动生成ID
    
    def add_item(self):
        """添加新物资（入库）"""
        self.open_add_item_dialog('入库')
    
    def add_quantity(self, operation_type):
        """增加物资数量"""
        if self.current_view == 'operations':
            # 从操作记录视图选择物品
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning('提示', '请先选择物资')
                return
                
            idx = self.tree.index(selected[0])
            item = self.data[idx]
            item_id = item.get('物资编号', '')
            
            # 打开对话框
            self.open_operation_dialog(operation_type, item_id)
        else:
            # 从库存视图选择物品
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning('提示', '请先选择物资')
                return
                
            item_values = self.tree.item(selected[0], 'values')
            item_id = item_values[0]  # 第一列是物资编号
            
            # 打开对话框
            self.open_operation_dialog(operation_type, item_id)
    
    def open_operation_dialog(self, operation_type, item_id):
        """打开操作对话框，用于物资增添或部分出库"""
        win = tk.Toplevel(self.root)
        win.title(operation_type)
        win.geometry('300x350')
        
        # 获取物品信息
        current_qty = 0
        item_name = ""
        
        if item_id in self.inventory:
            inventory_item = self.inventory[item_id]
            current_qty = inventory_item.get('物品数量', 0)
            item_name = inventory_item.get('物品名称', '')
        else:
            # 如果在库存中找不到物品
            messagebox.showwarning('提示', f'在当前库存中找不到编号为"{item_id}"的物品')
            win.destroy()
            return
        
        tk.Label(win, text=f'当前物品: {item_name}').grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='w')
        tk.Label(win, text=f'当前数量: {current_qty}').grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='w')
        
        # 物资操作字段
        tk.Label(win, text='物资操作').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        operation_var = tk.StringVar(value=operation_type)
        operation_dropdown = ttk.Combobox(win, textvariable=operation_var, 
                                         values=['入库', '出库', '物资增添', '部分出库'], 
                                         state="readonly")
        operation_dropdown.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        
        # 数量输入
        tk.Label(win, text='操作数量').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        qty_entry = tk.Entry(win)
        qty_entry.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        if operation_type == '物资增添':
            qty_entry.insert(0, '1')  # 默认为1
        
        # 时间输入
        tk.Label(win, text='时间').grid(row=4, column=0, padx=5, pady=5, sticky='w')
        time_entry = tk.Entry(win)
        time_entry.grid(row=4, column=1, padx=5, pady=5, sticky='ew')
        time_entry.insert(0, datetime.datetime.now().strftime('%Y-%m-%d %H:%M'))
        
        # 操作人输入
        tk.Label(win, text='操作人').grid(row=5, column=0, padx=5, pady=5, sticky='w')
        operator_var = tk.StringVar()
        operator_entry = ttk.Combobox(win, textvariable=operator_var, values=self.operators)
        operator_entry.grid(row=5, column=1, padx=5, pady=5, sticky='ew')
        
        # 提交者输入
        tk.Label(win, text='提交者').grid(row=6, column=0, padx=5, pady=5, sticky='w')
        submitter_var = tk.StringVar()
        submitter_entry = ttk.Combobox(win, textvariable=submitter_var, values=self.operators)
        submitter_entry.grid(row=6, column=1, padx=5, pady=5, sticky='ew')
        
        # 保存按钮
        tk.Button(win, text='保存', 
                 command=lambda: self.save_operation(
                     win, operation_var.get(), item_id, qty_entry, time_entry,
                     operator_var, submitter_var, current_qty
                 )).grid(row=7, column=0, columnspan=2, pady=10)
    
    def save_operation(self, win, operation_type, item_id, qty_entry, time_entry, operator_var, submitter_var, current_qty):
        """保存操作结果"""
        try:
            # 获取操作数量
            try:
                qty = int(qty_entry.get())
                if qty <= 0:
                    raise ValueError('操作数量必须大于0')
            except ValueError:
                raise ValueError('请输入有效的数量')
            
            # 获取时间
            time_str = time_entry.get()
            try:
                datetime.datetime.strptime(time_str, '%Y-%m-%d %H:%M')
            except ValueError:
                raise ValueError('时间格式不正确，应为：年-月-日 时:分 (如 2023-05-16 14:30)')
            
            # 获取操作人和提交者
            operator = operator_var.get().strip()
            submitter = submitter_var.get().strip()
            
            if not operator:
                raise ValueError('请输入操作人')
            if not submitter:
                raise ValueError('请输入提交者')
            
            # 获取物品信息
            inventory_item = self.inventory.get(item_id, {})
            item_name = inventory_item.get('物品名称', '')
            organization = inventory_item.get('所属组织', '')
            
            # 创建操作记录
            operation = {
                "提交时间": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "物资编号": item_id,
                "物品名称": item_name,
                "物资操作": operation_type,
                "所属组织": organization,
                "物品数量": qty,
                "时间": time_str,
                "操作人": operator,
                "提交者": submitter
            }
            
            # 添加操作记录
            self.data.append(operation)
            self.save_data()
            
            # 更新库存
            if operation_type == '物资增添':
                if item_id in self.inventory:
                    self.inventory[item_id]['物品数量'] += qty
                    self.inventory[item_id]['最后操作'] = operation_type
                    self.inventory[item_id]['最后操作人'] = operator
                    self.inventory[item_id]['最后操作时间'] = time_str
                    new_qty = self.inventory[item_id]['物品数量']
                    messagebox.showinfo('成功', f'已增加 {qty} 个物品，现有 {new_qty} 个')
                else:
                    messagebox.showerror('错误', f'库存中不存在编号为"{item_id}"的物品')
            
            elif operation_type == '部分出库':
                if item_id in self.inventory:
                    if qty > self.inventory[item_id]['物品数量']:
                        raise ValueError(f'出库数量不能超过当前库存 ({self.inventory[item_id]["物品数量"]})')
                    
                    self.inventory[item_id]['物品数量'] -= qty
                    self.inventory[item_id]['最后操作'] = operation_type
                    self.inventory[item_id]['最后操作人'] = operator
                    self.inventory[item_id]['最后操作时间'] = time_str
                    
                    new_qty = self.inventory[item_id]['物品数量']
                    
                    if new_qty <= 0:
                        # 如果数量减至0或以下，移除物品
                        del self.inventory[item_id]
                        messagebox.showinfo('成功', f'已出库 {qty} 个物品，物品已从库存中移除')
                    else:
                        messagebox.showinfo('成功', f'已出库 {qty} 个物品，剩余 {new_qty} 个')
                else:
                    messagebox.showerror('错误', f'库存中不存在编号为"{item_id}"的物品')
            
            self.save_inventory()
            self.update_table()
            win.destroy()
            
            # 更新操作者和提交者到配置
            self.update_operators([operator, submitter])
            
        except Exception as e:
            messagebox.showerror('错误', str(e))
    
    def open_add_item_dialog(self, operation_type):
        """打开添加物资对话框（入库）"""
        win = tk.Toplevel(self.root)
        win.title(operation_type)
        win.geometry('400x400')
        
        # 创建基本字段
        tk.Label(win, text='物资编号').grid(row=0, column=0, padx=5, pady=5, sticky='w')
        entry_id = tk.Entry(win)
        entry_id.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        
        # 添加标签提示
        tk.Label(win, text='例如: A1-3-05, B2-5-13', fg='gray').grid(row=0, column=2, padx=5, pady=5, sticky='w')
        
        tk.Label(win, text='物品名称').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        entry_name = tk.Entry(win)
        entry_name.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        # 添加物资操作字段
        tk.Label(win, text='物资操作').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        operation_var = tk.StringVar(value=operation_type)
        operation_dropdown = ttk.Combobox(win, textvariable=operation_var, 
                                         values=['入库', '出库', '物资增添', '部分出库'], 
                                         state="readonly")
        operation_dropdown.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        
        tk.Label(win, text='所属组织').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        # 使用下拉列表选择组织
        org_var = tk.StringVar()
        if self.organizations:
            org_var.set(self.organizations[0])
        org_dropdown = ttk.Combobox(win, textvariable=org_var, values=self.organizations, state="readonly")
        org_dropdown.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        
        tk.Label(win, text='物品数量').grid(row=4, column=0, padx=5, pady=5, sticky='w')
        entry_count = tk.Entry(win)
        entry_count.grid(row=4, column=1, padx=5, pady=5, sticky='ew')
        
        tk.Label(win, text='时间').grid(row=5, column=0, padx=5, pady=5, sticky='w')
        entry_date = tk.Entry(win)
        entry_date.grid(row=5, column=1, padx=5, pady=5, sticky='ew')
        
        # 添加操作人输入
        tk.Label(win, text='操作人').grid(row=6, column=0, padx=5, pady=5, sticky='w')
        operator_var = tk.StringVar()
        operator_entry = ttk.Combobox(win, textvariable=operator_var, values=self.operators)
        operator_entry.grid(row=6, column=1, padx=5, pady=5, sticky='ew')
        
        # 添加提交者输入
        tk.Label(win, text='提交者').grid(row=7, column=0, padx=5, pady=5, sticky='w')
        submitter_var = tk.StringVar()
        submitter_entry = ttk.Combobox(win, textvariable=submitter_var, values=self.operators)
        submitter_entry.grid(row=7, column=1, padx=5, pady=5, sticky='ew')
        
        # 不再自动生成编号
        
        # 默认填入当前日期时间
        now = datetime.datetime.now()
        entry_date.insert(0, now.strftime('%Y-%m-%d %H:%M'))
        
        # 保存按钮
        tk.Button(win, text='保存', 
                 command=lambda: self.save_new_item(
                     win, entry_id, entry_name, operation_var, org_var, entry_count, entry_date,
                     operator_var, submitter_var
                 )).grid(row=8, column=0, columnspan=2, pady=10)
    
    def save_new_item(self, win, entry_id, entry_name, operation_var, org_var, entry_count, entry_date, operator_var, submitter_var):
        """保存新添加的物资"""
        try:
            item_id = entry_id.get().strip()
            item_name = entry_name.get().strip()
            operation = operation_var.get()
            organization = org_var.get()
            operator = operator_var.get().strip()
            submitter = submitter_var.get().strip()
            
            # 验证编号是否为空
            if not item_id:
                raise ValueError('请输入物资编号')
            
            # 检查编号是否重复
            if operation == '入库' and item_id in self.inventory:
                raise ValueError('编号已存在于库存中，请使用其他编号或选择"物资增添"操作')
            
            # 验证操作者和提交者是否填写
            if not operator:
                raise ValueError('请输入操作人姓名')
                
            if not submitter:
                raise ValueError('请输入提交者姓名')
                
            try:
                count = int(entry_count.get())
                if count <= 0:
                    raise ValueError('数量必须大于0')
            except:
                raise ValueError('请输入有效的数量')
            
            now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            item = {
                "提交时间": now,
                "物资编号": item_id,
                "物品名称": item_name,
                "物资操作": operation,
                "所属组织": organization,
                "物品数量": count,
                "时间": entry_date.get(),
                "操作人": operator,
                "提交者": submitter
            }
            
            # 验证所有字段
            if not item['物资编号'] or not item['物品名称'] or not item['所属组织'] or not item['时间']:
                raise ValueError('编号、名称、所属组织和时间为必填项')
                
            # 验证日期时间格式
            time_str = item['时间']
            try:
                datetime.datetime.strptime(time_str, '%Y-%m-%d %H:%M')
            except ValueError:
                raise ValueError('时间格式不正确，应为：年-月-日 时:分 (如 2023-05-16 14:30)')
            
            # 添加到操作记录
            self.data.append(item)
            self.save_data()
            
            # 更新库存
            if operation == '入库':
                # 更新或添加库存
                self.inventory[item_id] = {
                    "物资编号": item_id,
                    "物品名称": item_name,
                    "所属组织": organization,
                    "物品数量": count,
                    "最后操作": operation,
                    "最后操作人": operator,
                    "最后操作时间": time_str,
                    "备注": ""
                }
                self.save_inventory()
                
            self.update_table()
            win.destroy()
            
            # 更新操作者和提交者到配置
            self.update_operators([operator, submitter])
            
        except Exception as e:
            messagebox.showerror('错误', str(e))
    
    def open_complete_removal_dialog(self, item_id):
        """打开完全出库对话框"""
        win = tk.Toplevel(self.root)
        win.title('完全出库')
        win.geometry('300x250')
        
        # 获取物品信息
        if item_id in self.inventory:
            inventory_item = self.inventory[item_id]
            item_name = inventory_item.get('物品名称', '')
        else:
            messagebox.showerror('错误', f'库存中不存在编号为"{item_id}"的物品')
            win.destroy()
            return
        
        tk.Label(win, text=f'物品: {item_name}').grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='w')
        
        # 物资操作字段
        tk.Label(win, text='物资操作').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        operation_var = tk.StringVar(value='出库')
        operation_dropdown = ttk.Combobox(win, textvariable=operation_var, 
                                         values=['出库'], 
                                         state="readonly")
        operation_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        # 时间输入
        tk.Label(win, text='时间').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        time_entry = tk.Entry(win)
        time_entry.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        time_entry.insert(0, datetime.datetime.now().strftime('%Y-%m-%d %H:%M'))
        
        # 操作人输入
        tk.Label(win, text='操作人').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        operator_var = tk.StringVar()
        operator_entry = ttk.Combobox(win, textvariable=operator_var, values=self.operators)
        operator_entry.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        
        # 提交者输入
        tk.Label(win, text='提交者').grid(row=4, column=0, padx=5, pady=5, sticky='w')
        submitter_var = tk.StringVar()
        submitter_entry = ttk.Combobox(win, textvariable=submitter_var, values=self.operators)
        submitter_entry.grid(row=4, column=1, padx=5, pady=5, sticky='ew')
        
        # 确认按钮
        tk.Button(win, text='确认出库', 
                 command=lambda: self.complete_item_removal(
                     win, item_id, time_entry, operator_var, submitter_var
                 )).grid(row=5, column=0, columnspan=2, pady=10)
    
    def complete_item_removal(self, win, item_id, time_entry, operator_var, submitter_var):
        """完成物品完全出库"""
        try:
            # 获取时间
            time_str = time_entry.get()
            try:
                datetime.datetime.strptime(time_str, '%Y-%m-%d %H:%M')
            except ValueError:
                raise ValueError('时间格式不正确，应为：年-月-日 时:分 (如 2023-05-16 14:30)')
                
            operator = operator_var.get().strip()
            submitter = submitter_var.get().strip()
            
            if not operator:
                raise ValueError('请输入操作人')
            if not submitter:
                raise ValueError('请输入提交者')
            
            # 获取物品信息
            if item_id not in self.inventory:
                raise ValueError(f'库存中不存在编号为"{item_id}"的物品')
                
            inventory_item = self.inventory[item_id]
            item_name = inventory_item.get('物品名称', '')
            organization = inventory_item.get('所属组织', '')
            qty = inventory_item.get('物品数量', 0)
            
            # 创建操作记录
            operation = {
                "提交时间": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "物资编号": item_id,
                "物品名称": item_name,
                "物资操作": '出库',
                "所属组织": organization,
                "物品数量": qty,
                "时间": time_str,
                "操作人": operator,
                "提交者": submitter
            }
            
            # 添加操作记录
            self.data.append(operation)
            self.save_data()
            
            # 记录日志
            self.log_operation_to_file(operation, operator, submitter, "完全出库")
            
            # 从库存中删除物品
            del self.inventory[item_id]
            self.save_inventory()
            
            self.update_table()
            
            messagebox.showinfo('出库成功', '物资已完全出库！')
            win.destroy()
            
            # 更新操作者和提交者到配置
            self.update_operators([operator, submitter])
            
        except Exception as e:
            messagebox.showerror('错误', str(e))
    
    def log_operation_to_file(self, item, operator, submitter, operation_type):
        """记录操作到历史文件"""
        now = datetime.datetime.now()
        log_dir = os.path.join(self.data_dir, 'logs')
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
            
        log_file = os.path.join(log_dir, f'operation_log_{now.strftime("%Y%m")}.txt')
        
        log_entry = (f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] "
                    f"操作: {operation_type}, 操作人: {operator}, 提交者: {submitter}, "
                    f"物品: {item.get('物品名称', '')}(编号:{item.get('物资编号', '')}), "
                    f"数量: {item.get('物品数量', 0)}\n")
        
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(log_entry)

    def sort_by(self, col, reverse):
        """按列排序表格数据"""
        if self.current_view == 'operations':
            # 操作记录视图排序
            column_map = {
                '物资编号': '物资编号', 
                '物品名称': '物品名称',
                '物资操作': '物资操作', 
                '所属组织': '所属组织', 
                '物品数量': '物品数量', 
                '时间': '时间',
                '操作人': '操作人',
                '提交者': '提交者',
                '提交时间': '提交时间'
            }
            
            if col in column_map:
                # 按照指定列排序
                key = column_map[col]
                self.data.sort(key=lambda x: x.get(key, ''), reverse=reverse)
                self.update_table()
                # 下次点击反向排序
                self.tree.heading(col, command=lambda: self.sort_by(col, not reverse))
        else:
            # 库存视图排序
            column_map = {
                '物资编号': '物资编号', 
                '物品名称': '物品名称',
                '所属组织': '所属组织', 
                '物品数量': '物品数量', 
                '最后操作': '最后操作',
                '最后操作人': '最后操作人',
                '最后操作时间': '最后操作时间',
                '备注': '备注'
            }
            
            if col in column_map:
                # 将字典转为列表便于排序
                inventory_list = list(self.inventory.values())
                
                # 按照指定列排序
                key = column_map[col]
                inventory_list.sort(key=lambda x: x.get(key, ''), reverse=reverse)
                
                # 更新排序后的库存字典
                self.inventory = {item['物资编号']: item for item in inventory_list}
                
                self.update_table()
                # 下次点击反向排序
                self.tree.heading(col, command=lambda: self.sort_by(col, not reverse))
    
    def import_excel(self):
        """从Excel导入数据"""
        file_path = filedialog.askopenfilename(
            filetypes=[('Excel文件', '*.xlsx *.xls')],
            title='选择要导入的Excel文件'
        )
        
        if not file_path:
            return
            
        try:
            # 打开Excel文件
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 获取表头行
            headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
            
            # 通过相似度匹配表头
            required_headers = ['物资编号', '物品名称', '物资操作', '所属组织', '物品数量', '时间', '操作人', '提交者']
            header_mapping = self.match_headers(headers, required_headers)
            
            missing_headers = [h for h in required_headers if h not in header_mapping]
            
            if missing_headers:
                messagebox.showerror('导入错误', f'Excel文件缺少必要的列: {", ".join(missing_headers)}')
                return
                
            # 读取数据
            new_items = []
            invalid_rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    # 获取单元格值
                    item_id = str(row[header_mapping['物资编号']].value or '').strip()
                    
                    # 读取基本信息
                    item_name = str(row[header_mapping['物品名称']].value or '')
                    operation = str(row[header_mapping['物资操作']].value or '入库')
                    organization = str(row[header_mapping['所属组织']].value or '')
                    
                    # 读取数量并验证
                    qty_cell = row[header_mapping['物品数量']].value
                    try:
                        quantity = int(qty_cell)
                        if quantity <= 0:
                            raise ValueError('数量必须大于0')
                    except (ValueError, TypeError):
                        invalid_rows.append(f'第{row_idx}行: 无效的数量')
                        continue
                    
                    # 读取时间
                    time_cell = row[header_mapping['时间']].value
                    if isinstance(time_cell, datetime.datetime):
                        item_time = time_cell.strftime('%Y-%m-%d %H:%M')
                    elif isinstance(time_cell, str):
                        try:
                            datetime.datetime.strptime(time_cell, '%Y-%m-%d %H:%M')
                            item_time = time_cell
                        except ValueError:
                            invalid_rows.append(f'第{row_idx}行: 时间格式不正确')
                            continue
                    else:
                        item_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    
                    # 操作人和提交者
                    operator = str(row[header_mapping['操作人']].value or '')
                    submitter = str(row[header_mapping['提交者']].value or '')
                    
                    # 验证必填字段
                    if not (item_id and item_name and organization):
                        invalid_rows.append(f'第{row_idx}行: 缺少必填字段')
                        continue
                    
                    # 验证编号是否为空
                    if not item_id:
                        invalid_rows.append(f'第{row_idx}行: 物资编号不能为空')
                        continue
                    
                    # 创建物资记录
                    new_item = {
                        "提交时间": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        "物资编号": item_id,
                        "物品名称": item_name,
                        "物资操作": operation,
                        "所属组织": organization,
                        "物品数量": quantity,
                        "时间": item_time,
                        "操作人": operator,
                        "提交者": submitter
                    }
                    
                    new_items.append(new_item)
                    
                except Exception as e:
                    invalid_rows.append(f'第{row_idx}行: {str(e)}')
            
            if invalid_rows:
                messagebox.showwarning('导入警告', 
                                      f'有{len(invalid_rows)}行数据格式不正确，已跳过:\n' + 
                                      '\n'.join(invalid_rows[:10]) +
                                      (f'\n...等共{len(invalid_rows)}个错误' if len(invalid_rows) > 10 else ''))
            
            if new_items:
                # 检查编号重复
                existing_ids = {item.get('物资编号') for item in self.data}
                duplicates = [item for item in new_items if item['物资编号'] in existing_ids]
                
                if duplicates:
                    if messagebox.askyesno('编号重复', 
                                         f'有{len(duplicates)}个物资编号与现有物资重复，是否覆盖现有数据？'):
                        # 删除重复的物资
                        dup_ids = {item['物资编号'] for item in duplicates}
                        self.data = [item for item in self.data if item.get('物资编号') not in dup_ids]
                    else:
                        # 不覆盖，只保留不重复的
                        new_items = [item for item in new_items if item['物资编号'] not in existing_ids]
                
                # 添加新物资
                self.data.extend(new_items)
                self.save_data()
                self.update_table()
                
                # 更新操作人和提交者列表
                operators = set()
                for item in new_items:
                    if item.get('操作人'):
                        operators.add(item['操作人'])
                    if item.get('提交者'):
                        operators.add(item['提交者'])
                
                self.update_operators(operators)
                
                messagebox.showinfo('导入成功', f'成功导入{len(new_items)}个物资记录')
            else:
                messagebox.showinfo('导入结果', '没有有效的物资记录被导入')
                
        except Exception as e:
            messagebox.showerror('导入错误', f'导入Excel时发生错误: {str(e)}')
    
    def export_excel(self):
        """导出数据为Excel文件"""
        if self.current_view == 'operations' and not self.data:
            messagebox.showinfo('提示', '没有操作记录可导出')
            return
        elif self.current_view == 'inventory' and not self.inventory:
            messagebox.showinfo('提示', '没有库存数据可导出')
            return
            
        # 设置默认保存位置和文件名
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        if self.current_view == 'operations':
            default_filename = f"仓库操作记录_{timestamp}"
        else:
            default_filename = f"仓库库存状态_{timestamp}"
        
        file_path = filedialog.asksaveasfilename(
            initialdir=self.output_dir,
            initialfile=default_filename,
            defaultextension='.xlsx', 
            filetypes=[('Excel文件', '*.xlsx')]
        )
        
        if not file_path:
            return
            
        # 创建Excel文件
        self.create_excel_file(file_path)
    
    def create_excel_file(self, file_path):
        """创建Excel文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if self.current_view == 'operations':
            # 导出操作记录
            ws.title = "仓库操作记录"
            
            # 添加表头
            headers = ['提交时间', '物资编号', '物品名称', '物资操作', '所属组织', '物品数量', '时间', '操作人', '提交者']
            ws.append(headers)
            
            for item in self.data:
                row_data = [
                    item.get('提交时间', ''),
                    item.get('物资编号', ''),
                    item.get('物品名称', ''),
                    item.get('物资操作', ''),
                    item.get('所属组织', ''),
                    item.get('物品数量', 0),
                    item.get('时间', ''),
                    item.get('操作人', ''),
                    item.get('提交者', '')
                ]
                ws.append(row_data)
        else:
            # 导出库存状态
            ws.title = "仓库库存状态"
            
            # 添加表头
            headers = ['物资编号', '物品名称', '所属组织', '物品数量', '最后操作', '最后操作人', '最后操作时间', '备注']
            ws.append(headers)
            
            for item_id, item in self.inventory.items():
                row_data = [
                    item_id,
                    item.get('物品名称', ''),
                    item.get('所属组织', ''),
                    item.get('物品数量', 0),
                    item.get('最后操作', ''),
                    item.get('最后操作人', ''),
                    item.get('最后操作时间', ''),
                    item.get('备注', '')
                ]
                ws.append(row_data)
        
        wb.save(file_path)
        messagebox.showinfo('导出成功', f'数据已导出到 {file_path}')
    
    def match_headers(self, actual_headers, required_headers):
        """匹配表头，返回匹配的列索引映射
        
        Args:
            actual_headers: 实际的Excel表头列表
            required_headers: 需要的表头列表
            
        Returns:
            字典 {需要的表头: 对应的列索引}
        """
        header_mapping = {}
        
        # 清理表头（去除括号和其中的内容）
        cleaned_headers = [self.clean_header(header) for header in actual_headers]
        
        # 对每个需要的表头，找到最匹配的实际表头
        for req_header in required_headers:
            best_match = None
            best_score = -1
            
            for idx, (raw_header, clean_header) in enumerate(zip(actual_headers, cleaned_headers)):
                # 如果完全匹配（清理后）
                if clean_header == req_header:
                    best_match = idx
                    break
                
                # 简单相似度评分：包含关系
                if req_header in clean_header:
                    score = len(req_header) / len(clean_header) if clean_header else 0
                    if score > best_score:
                        best_score = score
                        best_match = idx
            
            # 如果找到匹配
            if best_match is not None:
                header_mapping[req_header] = best_match
        
        return header_mapping
    
    def clean_header(self, header):
        """清理表头，去除括号及其内容"""
        if not header:
            return ""
        # 去除（...）内容
        import re
        return re.sub(r'（.*?）', '', header).strip()

    def remove_item(self, operation_type):
        """移除物资（出库或部分出库）"""
        if self.current_view == 'operations':
            # 从操作记录视图选择物品
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning('提示', '请先选择物资')
                return
                
            idx = self.tree.index(selected[0])
            item = self.data[idx]
            item_id = item.get('物资编号', '')
            
            # 检查物品是否存在于库存
            if item_id not in self.inventory:
                messagebox.showwarning('提示', f'物品编号"{item_id}"在当前库存中不存在')
                return
                
            if operation_type == '出库':
                self.open_complete_removal_dialog(item_id)
            else:  # 部分出库
                self.open_operation_dialog(operation_type, item_id)
        else:
            # 从库存视图选择物品
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning('提示', '请先选择物资')
                return
                
            item_values = self.tree.item(selected[0], 'values')
            item_id = item_values[0]  # 第一列是物资编号
            
            if operation_type == '出库':
                self.open_complete_removal_dialog(item_id)
            else:  # 部分出库
                self.open_operation_dialog(operation_type, item_id)
    
    def update_operators(self, new_operators):
        """更新操作者列表并保存到配置"""
        # 转换为集合以去重
        operators_set = set(self.operators)
        
        # 添加新操作者
        for operator in new_operators:
            if operator and operator not in operators_set:
                self.operators.append(operator)
                operators_set.add(operator)
        
        # 保存到配置文件
        self.save_config()

    def save_config(self):
        """保存配置到文件"""
        config = {
            "organization": {
                "val": self.organizations
            },
            "operators": {
                "val": self.operators
            }
        }
        
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror('配置保存错误', f'无法保存配置: {str(e)}')
    
if __name__ == '__main__':
    root = tk.Tk()
    app = WarehouseManager(root)
    root.mainloop()
